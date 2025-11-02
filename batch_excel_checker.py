#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
סוכן בדיקת מטלות אקסל אקדמיות - גרסה מתקדמת עם בדיקות מרובות
תומך ב-Streamlit Cloud, GitHub, ו-Claude API
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import pandas as pd
import json
import sys
import re
from datetime import datetime
from pathlib import Path
from typing import Dict, List, Tuple, Any
from difflib import SequenceMatcher
import anthropic
import os


class BatchExcelChecker:
    """מחלקה לבדיקת מספר מטלות אקסל בבת אחת"""
    
    def __init__(self, rubric_file: str, config: Dict = None, 
                 output_dir: str = "results", use_ai: bool = False):
        """
        אתחול הבודק
        
        Args:
            rubric_file: נתיב לקובץ המחוון
            config: הגדרות מתקדמות (אופציונלי)
            output_dir: תיקייה לשמירת התוצאות
            use_ai: האם להשתמש ב-Claude API לבדיקות מתקדמות
        """
        self.rubric_file = rubric_file
        self.output_dir = Path(output_dir)
        self.output_dir.mkdir(exist_ok=True)
        self.use_ai = use_ai
        
        # הגדרות ברירת מחדל
        self.config = {
            'sheet_name_similarity_threshold': 0.6,
            'check_formulas': True,
            'check_functions': True,
            'check_references': True,
            'partial_credit': True,
            'strict_mode': False
        }
        
        if config:
            self.config.update(config)
        
        # אתחול Claude API אם נדרש
        self.claude_client = None
        if use_ai and os.getenv('ANTHROPIC_API_KEY'):
            self.claude_client = anthropic.Anthropic(
                api_key=os.getenv('ANTHROPIC_API_KEY')
            )
        
        self.rubric_wb = None
        self.batch_results = []
        self.summary_df = None
    
    def load_rubric(self) -> bool:
        """טעינת קובץ המחוון"""
        try:
            print(f"📁 טוען קובץ מחוון: {self.rubric_file}")
            self.rubric_wb = openpyxl.load_workbook(self.rubric_file, data_only=True)
            print("✓ מחוון נטען בהצלחה\n")
            return True
        except Exception as e:
            print(f"✗ שגיאה בטעינת מחוון: {str(e)}")
            return False
    
    def check_single_student(self, student_file: str, student_id: str = None) -> Dict:
        """
        בדיקת מטלה של תלמיד בודד
        
        Args:
            student_file: נתיב לקובץ התלמיד
            student_id: מזהה התלמיד (אופציונלי)
        
        Returns:
            תוצאות הבדיקה
        """
        from excel_checker_advanced import AdvancedExcelChecker
        
        if not student_id:
            student_id = Path(student_file).stem
        
        print(f"\n{'='*80}")
        print(f"🔍 בודק מטלה: {student_id}")
        print(f"{'='*80}")
        
        # יצירת בודק למטלה זו
        checker = AdvancedExcelChecker(
            self.rubric_file,
            student_file,
            self.config,
            output_dir=str(self.output_dir / student_id)
        )
        
        # הרצת הבדיקה
        if checker.run_checks():
            results = checker.results
            results['student_id'] = student_id
            results['student_file'] = student_file
            
            # שמירת תוצאות
            self.batch_results.append(results)
            
            # הוספת גליון בדיקה לקובץ המטלה
            self._add_grading_sheet_to_file(student_file, results, student_id)
            
            return results
        
        return None
    
    def _add_grading_sheet_to_file(self, student_file: str, results: Dict, student_id: str):
        """
        הוספת גליון בדיקה לקובץ המטלה המקורי
        
        Args:
            student_file: נתיב לקובץ התלמיד
            results: תוצאות הבדיקה
            student_id: מזהה התלמיד
        """
        try:
            # טעינת הקובץ
            wb = openpyxl.load_workbook(student_file)
            
            # יצירת גליון חדש
            ws = wb.create_sheet("🎓 גליון_בדיקה", 0)
            
            # עיצוב כותרת
            ws.merge_cells('A1:F1')
            title_cell = ws['A1']
            title_cell.value = "דוח בדיקת מטלה אוטומטית"
            title_cell.font = Font(size=16, bold=True, color="FFFFFF")
            title_cell.fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
            title_cell.alignment = Alignment(horizontal='center', vertical='center')
            ws.row_dimensions[1].height = 30
            
            # פרטי תלמיד
            row = 3
            ws[f'A{row}'] = "מספר מטלה:"
            ws[f'B{row}'] = student_id
            ws[f'A{row}'].font = Font(bold=True)
            
            row += 1
            ws[f'A{row}'] = "תאריך בדיקה:"
            ws[f'B{row}'] = datetime.now().strftime("%d/%m/%Y %H:%M")
            ws[f'A{row}'].font = Font(bold=True)
            
            # ציון כללי
            row += 2
            ws.merge_cells(f'A{row}:F{row}')
            ws[f'A{row}'] = "סיכום ציונים"
            ws[f'A{row}'].font = Font(size=14, bold=True)
            ws[f'A{row}'].fill = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
            
            row += 1
            ws[f'A{row}'] = "ציון כולל:"
            ws[f'B{row}'] = f"{results['total_score']:.1f} / {results['max_score']}"
            ws[f'A{row}'].font = Font(bold=True)
            ws[f'B{row}'].font = Font(size=12, bold=True, color="0000FF")
            
            row += 1
            ws[f'A{row}'] = "אחוז:"
            ws[f'B{row}'] = f"{results['percentage']:.1f}%"
            ws[f'A{row}'].font = Font(bold=True)
            
            # צביעה לפי ציון
            percentage = results['percentage']
            if percentage >= 80:
                color = "00B050"  # ירוק
            elif percentage >= 60:
                color = "FFC000"  # כתום
            else:
                color = "FF0000"  # אדום
            ws[f'B{row}'].font = Font(size=12, bold=True, color=color)
            
            # כותרות טבלה
            row += 2
            headers = ['מס\'', 'סעיף', 'תת-סעיף', 'סטטוס', 'ציון', 'הערות']
            for col_idx, header in enumerate(headers, start=1):
                cell = ws.cell(row=row, column=col_idx, value=header)
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
                cell.alignment = Alignment(horizontal='center')
            
            # נתוני הבדיקות
            row += 1
            for idx, check in enumerate(results['checks'], start=1):
                ws.cell(row=row, column=1, value=idx)
                ws.cell(row=row, column=2, value=check['section'])
                ws.cell(row=row, column=3, value=check['subsection'])
                
                # סטטוס עם צבע
                status_cell = ws.cell(row=row, column=4, value=check['status'])
                if check['status'] == 'עבר':
                    status_cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                elif check['status'] == 'עבר חלקית':
                    status_cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
                else:
                    status_cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                
                ws.cell(row=row, column=5, value=f"{check['earned_points']:.1f}/{check['max_points']}")
                
                # הערות
                notes = '\n'.join(check['notes']) if check['notes'] else ''
                notes_cell = ws.cell(row=row, column=6, value=notes)
                notes_cell.alignment = Alignment(wrap_text=True, vertical='top')
                
                row += 1
            
            # עיצוב טבלה
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            for row_cells in ws.iter_rows(min_row=8, max_row=row-1, min_col=1, max_col=6):
                for cell in row_cells:
                    cell.border = thin_border
            
            # התאמת רוחב עמודות
            ws.column_dimensions['A'].width = 6
            ws.column_dimensions['B'].width = 25
            ws.column_dimensions['C'].width = 30
            ws.column_dimensions['D'].width = 15
            ws.column_dimensions['E'].width = 12
            ws.column_dimensions['F'].width = 50
            
            # שמירת הקובץ
            output_file = self.output_dir / student_id / f"{student_id}_עם_בדיקה.xlsx"
            output_file.parent.mkdir(parents=True, exist_ok=True)
            wb.save(output_file)
            
            print(f"✓ גליון בדיקה נוסף לקובץ: {output_file}")
            
        except Exception as e:
            print(f"⚠️  שגיאה בהוספת גליון בדיקה: {e}")
    
    def check_batch(self, student_files: List[str], student_ids: List[str] = None) -> bool:
        """
        בדיקת מספר מטלות בבת אחת
        
        Args:
            student_files: רשימת נתיבים לקבצי תלמידים
            student_ids: רשימת מזהי תלמידים (אופציונלי)
        
        Returns:
            האם הבדיקה הצליחה
        """
        if not self.load_rubric():
            return False
        
        if not student_ids:
            student_ids = [Path(f).stem for f in student_files]
        
        print(f"\n🎓 מתחיל בדיקת {len(student_files)} מטלות")
        print("="*80)
        
        for student_file, student_id in zip(student_files, student_ids):
            self.check_single_student(student_file, student_id)
        
        # יצירת קובץ סיכום
        self._create_summary_excel()
        
        print(f"\n{'='*80}")
        print(f"✅ בדיקת כל המטלות הושלמה!")
        print(f"📊 נבדקו {len(self.batch_results)} מטלות")
        print(f"📁 תוצאות נשמרו ב: {self.output_dir}")
        print(f"{'='*80}\n")
        
        return True
    
    def _create_summary_excel(self):
        """יצירת קובץ Excel סיכום לכל המטלות"""
        
        summary_data = []
        for result in self.batch_results:
            row = {
                'מספר_מטלה': result['student_id'],
                'ציון': result['total_score'],
                'מקסימום': result['max_score'],
                'אחוז': result['percentage'],
                'סטטוס': 'עבר' if result['percentage'] >= 80 else 'עבר חלקית' if result['percentage'] >= 60 else 'נכשל',
                'בדיקות_שעברו': result['summary']['passed'],
                'בדיקות_שנכשלו': result['summary']['failed'],
                'תאריך_בדיקה': result['check_date']
            }
            
            # הערות - מה ירד
            failed_checks = [c for c in result['checks'] if c['status'] != 'עבר']
            notes = []
            for check in failed_checks:
                deduction = check['max_points'] - check['earned_points']
                if deduction > 0:
                    notes.append(f"{check['section']}: -{deduction:.1f} נקודות")
            
            row['הערות_מה_ירד'] = '\n'.join(notes) if notes else 'הכל תקין'
            
            summary_data.append(row)
        
        # יצירת DataFrame
        df = pd.DataFrame(summary_data)
        
        # שמירה לExcel עם עיצוב
        summary_file = self.output_dir / f"סיכום_כל_המטלות_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        
        with pd.ExcelWriter(summary_file, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='סיכום', index=False)
            
            # עיצוב
            workbook = writer.book
            worksheet = writer.sheets['סיכום']
            
            # כותרות
            for cell in worksheet[1]:
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
                cell.alignment = Alignment(horizontal='center')
            
            # התאמת רוחב עמודות
            for column in worksheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                worksheet.column_dimensions[column_letter].width = adjusted_width
            
            # צביעת שורות לפי סטטוס
            for row in range(2, len(df) + 2):
                status_cell = worksheet.cell(row=row, column=5)
                if status_cell.value == 'עבר':
                    for col in range(1, len(df.columns) + 1):
                        worksheet.cell(row=row, column=col).fill = PatternFill(
                            start_color="E2EFDA", end_color="E2EFDA", fill_type="solid"
                        )
                elif status_cell.value == 'נכשל':
                    for col in range(1, len(df.columns) + 1):
                        worksheet.cell(row=row, column=col).fill = PatternFill(
                            start_color="FCE4D6", end_color="FCE4D6", fill_type="solid"
                        )
        
        self.summary_df = df
        print(f"\n📊 קובץ סיכום נוצר: {summary_file}")
        
        return summary_file


# דוגמת שימוש
if __name__ == "__main__":
    print("="*80)
    print("🎓 בודק מטלות אקסל - גרסת Batch")
    print("="*80)
    
    if len(sys.argv) < 3:
        print("\n📖 שימוש:")
        print("  python batch_excel_checker.py <מחוון> <תיקיית_מטלות>")
        print("\nדוגמה:")
        print("  python batch_excel_checker.py rubric.xlsx students/")
        sys.exit(1)
    
    rubric_file = sys.argv[1]
    students_dir = Path(sys.argv[2])
    
    # איסוף כל קבצי האקסל
    student_files = list(students_dir.glob("*.xlsx")) + list(students_dir.glob("*.xls"))
    student_files = [str(f) for f in student_files]
    
    if not student_files:
        print(f"❌ לא נמצאו קבצי אקסל בתיקייה: {students_dir}")
        sys.exit(1)
    
    print(f"\n📁 נמצאו {len(student_files)} מטלות לבדיקה")
    
    # יצירת הבודק
    checker = BatchExcelChecker(
        rubric_file=rubric_file,
        output_dir="batch_results",
        use_ai=False  # שנה ל-True אם יש API key
    )
    
    # הרצת בדיקת מטלות
    checker.check_batch(student_files)
