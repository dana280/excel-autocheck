#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
סוכן בדיקת מטלות אקסל אקדמיות - גרסה מתקדמת
תומך ב-Streamlit Cloud ו-GitHub
"""

import openpyxl
import pandas as pd
import json
import sys
import re
from datetime import datetime
from pathlib import Path
from typing import Dict, List, Tuple, Any
from difflib import SequenceMatcher


class AdvancedExcelChecker:
    """מחלקה מתקדמת לבדיקת מטלות אקסל"""
    
    def __init__(self, rubric_file: str, student_file: str, 
                 config: Dict = None, output_dir: str = "results"):
        """
        אתחול הבודק
        
        Args:
            rubric_file: נתיב לקובץ המחוון
            student_file: נתיב לקובץ התלמיד
            config: הגדרות מתקדמות (אופציונלי)
            output_dir: תיקייה לשמירת התוצאות
        """
        self.rubric_file = rubric_file
        self.student_file = student_file
        self.output_dir = Path(output_dir)
        self.output_dir.mkdir(exist_ok=True)
        
        # הגדרות ברירת מחדל
        self.config = {
            'sheet_name_similarity_threshold': 0.6,
            'check_formulas': True,
            'check_functions': True,
            'check_references': True,
            'partial_credit': True,
            'strict_mode': False
        }
        
        if config:
            self.config.update(config)
        
        self.rubric_wb = None
        self.student_wb = None
        self.results = {
            'student_file': student_file,
            'check_date': datetime.now().isoformat(),
            'total_score': 0,
            'max_score': 0,
            'percentage': 0,
            'checks': [],
            'sheet_mapping': {},
            'summary': {},
            'warnings': []
        }
    
    def load_files(self) -> bool:
        """טעינת קבצי האקסל"""
        try:
            print(f"📁 טוען קובץ מחוון: {self.rubric_file}")
            self.rubric_wb = openpyxl.load_workbook(self.rubric_file, data_only=True)
            
            print(f"📁 טוען קובץ תלמיד: {self.student_file}")
            # data_only=False כדי לקרוא נוסחאות
            self.student_wb = openpyxl.load_workbook(self.student_file, data_only=False)
            
            print("✓ קבצים נטענו בהצלחה\n")
            return True
        except Exception as e:
            print(f"✗ שגיאה בטעינת קבצים: {str(e)}")
            self.results['warnings'].append(f"שגיאה בטעינת קבצים: {str(e)}")
            return False
    
    def find_similar_sheet(self, target_sheet: str) -> str:
        """מציאת גליון דומה בקובץ התלמיד"""
        if target_sheet in self.student_wb.sheetnames:
            return target_sheet
        
        # חיפוש גליון דומה
        best_match = None
        best_ratio = 0
        
        for sheet_name in self.student_wb.sheetnames:
            ratio = SequenceMatcher(None, 
                                   target_sheet.lower(), 
                                   sheet_name.lower()).ratio()
            
            if ratio > best_ratio and ratio >= self.config['sheet_name_similarity_threshold']:
                best_ratio = ratio
                best_match = sheet_name
        
        return best_match
    
    def parse_rubric(self) -> List[Dict]:
        """פענוח המחוון לרשימת בדיקות"""
        rubric_checks = []
        
        # הנחה: המחוון בגליון הראשון
        ws = self.rubric_wb.worksheets[0]
        
        print("📋 מפענח מחוון:")
        print("-" * 80)
        
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if not row or not any(row):  # שורה ריקה
                continue
            
            sheet_name = str(row[0]) if row[0] else ""
            section = str(row[1]) if row[1] else ""
            subsection = str(row[2]) if row[2] else ""
            
            # דילוג על שורות ללא שם גליון
            if not sheet_name:
                continue
            
            # ניקוד
            points_value = row[3] if len(row) > 3 else 0
            try:
                if isinstance(points_value, str) and points_value.startswith('='):
                    # חישוב נוסחה פשוטה
                    deduction_col_idx = 4
                    deduction = float(row[deduction_col_idx]) if len(row) > deduction_col_idx and row[deduction_col_idx] else 0
                    points = eval(points_value[1:].replace(f'E{row_idx}', str(deduction)))
                else:
                    points = float(points_value) if points_value else 0
            except Exception as e:
                print(f"⚠️  שגיאה בחישוב ניקוד בשורה {row_idx}: {e}")
                points = 0
            
            deduction = row[4] if len(row) > 4 and row[4] else 0
            
            # מציאת גליון מתאים
            actual_sheet = self.find_similar_sheet(sheet_name)
            
            check = {
                'row': row_idx,
                'sheet_rubric': sheet_name,
                'sheet_actual': actual_sheet,
                'section': section,
                'subsection': subsection,
                'max_points': points,
                'deduction': deduction,
                'earned_points': 0,
                'status': 'ממתין',
                'notes': [],
                'formulas_found': []
            }
            
            rubric_checks.append(check)
            
            # עדכון מיפוי גליונות
            if actual_sheet and sheet_name != actual_sheet:
                self.results['sheet_mapping'][sheet_name] = actual_sheet
            
            status_icon = "✓" if actual_sheet else "✗"
            print(f"  {status_icon} {sheet_name} → {actual_sheet or 'לא נמצא'}")
            print(f"     {section} | {subsection}: {points} נקודות")
        
        print(f"\n✓ נמצאו {len(rubric_checks)} בדיקות במחוון")
        return rubric_checks
    
    def check_formulas_in_sheet(self, sheet_name: str) -> Dict[str, Any]:
        """בדיקת נוסחאות בגליון"""
        if not sheet_name or sheet_name not in self.student_wb.sheetnames:
            return {
                'formulas_count': 0,
                'formulas': [],
                'has_formulas': False,
                'functions_used': {}
            }
        
        ws = self.student_wb[sheet_name]
        formulas = []
        functions_used = {}
        
        for row in ws.iter_rows():
            for cell in row:
                if cell.value and isinstance(cell.value, str) and cell.value.startswith('='):
                    formula = cell.value
                    formulas.append({
                        'cell': cell.coordinate,
                        'formula': formula,
                        'sheet': sheet_name
                    })
                    
                    # זיהוי פונקציות
                    functions = re.findall(r'([A-Z]+)\(', formula.upper())
                    for func in functions:
                        functions_used[func] = functions_used.get(func, 0) + 1
        
        return {
            'formulas_count': len(formulas),
            'formulas': formulas,
            'has_formulas': len(formulas) > 0,
            'functions_used': functions_used
        }
    
    def analyze_check_requirements(self, check: Dict) -> List[str]:
        """ניתוח דרישות הבדיקה מתוך הטקסט"""
        requirements = []
        text = f"{check['section']} {check['subsection']}".lower()
        
        # זיהוי פונקציות נדרשות
        function_keywords = {
            'sum': ['סה"כ', 'סכום', 'sum'],
            'if': ['אם', 'תנאי', 'if'],
            'vlookup': ['חיפוש', 'vlookup', 'lookup'],
            'countif': ['ספירה', 'מספר', 'countif', 'count'],
            'sumif': ['סיכום תנאי', 'sumif']
        }
        
        for func, keywords in function_keywords.items():
            if any(keyword in text for keyword in keywords):
                requirements.append(f'function:{func.upper()}')
        
        # זיהוי דרישות נוספות
        if 'תא עזר' in text or 'תא עזר' in text:
            requirements.append('uses_helper_cell')
        
        if 'הפניה' in text or 'reference' in text:
            requirements.append('has_reference')
        
        if 'תרשים' in text or 'chart' in text:
            requirements.append('has_chart')
        
        return requirements
    
    def check_requirements(self, check: Dict, formula_data: Dict) -> Tuple[float, List[str]]:
        """בדיקת עמידה בדרישות"""
        requirements = self.analyze_check_requirements(check)
        notes = []
        earned_ratio = 0
        
        if not requirements:
            # אם אין דרישות ספציפיות, בדוק רק שיש נוסחאות
            if formula_data['has_formulas']:
                earned_ratio = 1.0
                notes.append(f"✓ נמצאו {formula_data['formulas_count']} נוסחאות")
            else:
                notes.append("✗ לא נמצאו נוסחאות")
            return earned_ratio, notes
        
        met_requirements = 0
        total_requirements = len(requirements)
        
        for req in requirements:
            if req.startswith('function:'):
                func_name = req.split(':')[1]
                if func_name in formula_data['functions_used']:
                    met_requirements += 1
                    count = formula_data['functions_used'][func_name]
                    notes.append(f"✓ שימוש בפונקציה {func_name} ({count} פעמים)")
                else:
                    notes.append(f"✗ חסרה פונקציה {func_name}")
            
            elif req == 'uses_helper_cell':
                # בדיקה אם יש תא עזר (כרגע פשוט בודק שיש נוסחאות)
                if formula_data['has_formulas']:
                    met_requirements += 1
                    notes.append("✓ שימוש בתאי עזר")
            
            elif req == 'has_reference':
                # בדיקה אם יש הפניות לתאים אחרים
                has_refs = any('!' in f['formula'] for f in formula_data['formulas'])
                if has_refs:
                    met_requirements += 1
                    notes.append("✓ יש הפניות בין גליונות")
                else:
                    notes.append("⚠️  לא נמצאו הפניות בין גליונות")
        
        if total_requirements > 0:
            earned_ratio = met_requirements / total_requirements
        else:
            earned_ratio = 1.0 if formula_data['has_formulas'] else 0
        
        return earned_ratio, notes
    
    def run_checks(self) -> bool:
        """הרצת כל הבדיקות"""
        print("\n" + "="*80)
        print("🔍 מתחיל בדיקת מטלה")
        print("="*80)
        
        if not self.load_files():
            return False
        
        # פענוח המחוון
        checks = self.parse_rubric()
        total_earned = 0
        total_max = 0
        
        print("\n🔬 מבצע בדיקות:")
        print("-" * 80)
        
        for check in checks:
            sheet_name = check['sheet_actual']
            max_points = check['max_points']
            total_max += max_points
            
            print(f"\n📌 {check['sheet_rubric']} | {check['section']}")
            print(f"   {check['subsection']}")
            
            if not sheet_name:
                check['status'] = 'נכשל'
                check['earned_points'] = 0
                check['notes'].append(f"✗ גליון '{check['sheet_rubric']}' לא נמצא")
                print(f"   ✗ גליון לא נמצא")
                continue
            
            # בדיקת נוסחאות בגליון
            formula_data = self.check_formulas_in_sheet(sheet_name)
            check['formulas_found'] = formula_data['formulas'][:5]  # רק 5 ראשונות לדוגמה
            
            # בדיקת עמידה בדרישות
            earned_ratio, notes = self.check_requirements(check, formula_data)
            check['notes'].extend(notes)
            
            # חישוב ניקוד
            if self.config['partial_credit']:
                check['earned_points'] = max_points * earned_ratio
            else:
                check['earned_points'] = max_points if earned_ratio == 1.0 else 0
            
            total_earned += check['earned_points']
            
            # עדכון סטטוס
            if earned_ratio >= 0.8:
                check['status'] = 'עבר'
                status_icon = "✅"
            elif earned_ratio >= 0.5:
                check['status'] = 'עבר חלקית'
                status_icon = "⚠️"
            else:
                check['status'] = 'נכשל'
                status_icon = "❌"
            
            print(f"   {status_icon} ציון: {check['earned_points']:.1f}/{max_points}")
            for note in notes:
                print(f"      {note}")
        
        # סיכום
        self.results['checks'] = checks
        self.results['total_score'] = round(total_earned, 1)
        self.results['max_score'] = total_max
        self.results['percentage'] = round((total_earned / total_max * 100) if total_max > 0 else 0, 1)
        
        # סטטיסטיקות
        passed = sum(1 for c in checks if c['status'] == 'עבר')
        partial = sum(1 for c in checks if c['status'] == 'עבר חלקית')
        failed = sum(1 for c in checks if c['status'] == 'נכשל')
        
        self.results['summary'] = {
            'total_checks': len(checks),
            'passed': passed,
            'partial': partial,
            'failed': failed
        }
        
        print("\n" + "="*80)
        print("📊 סיכום בדיקה")
        print("="*80)
        print(f"ציון כולל: {self.results['total_score']}/{self.results['max_score']} ({self.results['percentage']}%)")
        print(f"בדיקות שעברו: {passed}/{len(checks)}")
        print(f"בדיקות חלקיות: {partial}/{len(checks)}")
        print(f"בדיקות שנכשלו: {failed}/{len(checks)}")
        
        return True
    
    def generate_report(self) -> Tuple[str, str]:
        """יצירת דוחות בדיקה"""
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        
        # דוח JSON
        report_json = self.output_dir / f"report_{timestamp}.json"
        with open(report_json, 'w', encoding='utf-8') as f:
            json.dump(self.results, f, ensure_ascii=False, indent=2)
        
        # דוח טקסט מפורט
        report_txt = self.output_dir / f"report_{timestamp}.txt"
        with open(report_txt, 'w', encoding='utf-8') as f:
            f.write("="*80 + "\n")
            f.write("דוח בדיקת מטלה אוטומטית - מערכת בדיקת אקסל\n")
            f.write("="*80 + "\n\n")
            f.write(f"📁 קובץ נבדק: {self.results['student_file']}\n")
            f.write(f"📅 תאריך בדיקה: {self.results['check_date']}\n")
            f.write(f"📊 ציון כולל: {self.results['total_score']}/{self.results['max_score']}")
            f.write(f" ({self.results['percentage']}%)\n\n")
            
            # סיכום
            summary = self.results['summary']
            f.write("סיכום:\n")
            f.write("-"*80 + "\n")
            f.write(f"✅ בדיקות שעברו: {summary['passed']}\n")
            f.write(f"⚠️  בדיקות חלקיות: {summary['partial']}\n")
            f.write(f"❌ בדיקות שנכשלו: {summary['failed']}\n")
            f.write(f"📋 סה\"כ בדיקות: {summary['total_checks']}\n\n")
            
            # מיפוי גליונות
            if self.results['sheet_mapping']:
                f.write("מיפוי גליונות:\n")
                f.write("-"*80 + "\n")
                for rubric_sheet, actual_sheet in self.results['sheet_mapping'].items():
                    f.write(f"  {rubric_sheet} → {actual_sheet}\n")
                f.write("\n")
            
            # פירוט בדיקות
            f.write("פירוט בדיקות:\n")
            f.write("="*80 + "\n")
            
            for i, check in enumerate(self.results['checks'], 1):
                f.write(f"\n{i}. {check['sheet_rubric']} | {check['section']} | {check['subsection']}\n")
                f.write(f"   סטטוס: {check['status']}\n")
                f.write(f"   ציון: {check['earned_points']:.1f}/{check['max_points']}\n")
                
                if check['notes']:
                    f.write(f"   הערות:\n")
                    for note in check['notes']:
                        f.write(f"     • {note}\n")
                
                if check['formulas_found']:
                    f.write(f"   דוגמאות לנוסחאות שנמצאו:\n")
                    for formula in check['formulas_found'][:3]:
                        f.write(f"     • {formula['cell']}: {formula['formula']}\n")
        
        print(f"\n📄 דוחות נשמרו:")
        print(f"   • JSON: {report_json}")
        print(f"   • TEXT: {report_txt}")
        
        return str(report_json), str(report_txt)


def main():
    """פונקציה ראשית"""
    print("="*80)
    print("🎓 סוכן בדיקת מטלות אקסל - גרסה מתקדמת")
    print("="*80)
    
    if len(sys.argv) < 3:
        print("\n📖 שימוש: python excel_checker_advanced.py <קובץ_מחוון> <קובץ_תלמיד>")
        print("\nדוגמה:")
        print("  python excel_checker_advanced.py indicator_new.xlsx solution_new.xlsx")
        return
    
    rubric_file = sys.argv[1]
    student_file = sys.argv[2]
    
    # הגדרות אופציונליות
    config = {
        'sheet_name_similarity_threshold': 0.6,
        'partial_credit': True,
        'strict_mode': False
    }
    
    checker = AdvancedExcelChecker(rubric_file, student_file, config)
    
    if checker.run_checks():
        checker.generate_report()
        print("\n✅ בדיקה הושלמה בהצלחה!")
    else:
        print("\n❌ הבדיקה נכשלה")
        sys.exit(1)


if __name__ == "__main__":
    main()
